
from openpyxl.styles import PatternFill
import openpyxl
from openpyxl.utils import get_column_letter

book = openpyxl.Workbook()

book.create_sheet('Meus Computadores')

print(book.sheetnames)

computadores = book['Meus Computadores']
computadores.append(['Computadores','RAM','Preços'])
computadores.append(['--------','---------','--------'])
computadores.append(['Computador Central','64gb','R$ 10.000'])
computadores.append(['Computador Reserva','32gb','R$ 5.000'])
computadores.append(['Computador 1','16gb','R$ 2.500'])
computadores.append(['Computador 2','16gb','R$ 2.500'])
computadores.append(['Computador 3','16gb','R$ 2.500'])
computadores.append(['Computador 4','16gb','R$ 2.500'])

azul = "ADD8E6" 

blueFill = PatternFill(start_color=azul, end_color=azul, fill_type= 'solid')
computadores['A1'].fill = blueFill
computadores['B1'].fill = blueFill
computadores['C1'].fill = blueFill

for row in range(1,9):
    for col in range(1,6):                                    
            computadores.row_dimensions[row].height = 30
            col_letter = get_column_letter(col)
            computadores.column_dimensions[col_letter].width = 20

book.save('Computadores Locais.xlsx')
